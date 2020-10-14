# -*- coding: utf-8 -*-
from collections import defaultdict
from datetime import date as dt

from django.conf import settings
from django.contrib.auth.models import Group
from django.contrib.postgres.search import SearchVector
from django.db.models import CharField, Value
from django.db.models.functions import Concat, Cast
from django.utils import timezone
from openpyxl import load_workbook

from esapi.helper.opdata import ESOPDataHelper
from localinfo.models import Operator, Commune, DayType, HalfHour, TimePeriod, TransportMode, GlobalPermission, \
    CalendarInfo, FAQ, OPDictionary, TimePeriodDate, OPProgram


def _list_parser(list):
    return [{'value': x[0], 'item': x[1]} for x in list]


def _dict_parser(list):
    return {x[0]: x[1] for x in list}


def get_day_type_list_for_select_input(to_dict=False):
    parser = _list_parser
    if to_dict:
        parser = _dict_parser
    return parser(DayType.objects.values_list('esId', 'name'))


def get_operator_list_for_select_input(filter=None, to_dict=False):
    """
    :param filter: list of elasticsearch_id to be returned
    :return: list of dict {esId: elasticsearch_id, text: operator_name}
    """
    parser = _list_parser
    queryset = Operator.objects.values_list('esId', 'name')
    if filter:
        queryset = queryset.filter(esId__in=filter)
    if to_dict:
        parser = _dict_parser
    return parser(queryset)


def get_timeperiod_list_for_select_input(to_dict=False, filter_id=None):
    parser = _list_parser
    if to_dict:
        parser = _dict_parser

    queryset = TimePeriod.objects.all()
    if filter_id is not None:
        queryset = queryset.filter(date_id=filter_id)

    return parser(queryset.values_list('esId').annotate(
        name=Concat('authorityPeriodName', Value(' ('), 'initialTime', Value('-'), 'endTime', Value(')'),
                    output_field=CharField())))


def get_halfhour_list_for_select_input(to_dict=False, format='longName'):
    parser = _list_parser
    if to_dict:
        parser = _dict_parser
    return parser(HalfHour.objects.values_list('esId', format))


def get_commune_list_for_select_input(to_dict=False):
    parser = _list_parser
    if to_dict:
        parser = _dict_parser
    return parser(Commune.objects.values_list('esId', 'name'))


def get_transport_mode_list_for_select_input(to_dict=False):
    parser = _list_parser
    if to_dict:
        parser = _dict_parser
    return parser(TransportMode.objects.values_list('esId', 'name'))


def get_calendar_info():
    """
    :return: all calendar info
    """
    query = CalendarInfo.objects.select_related('day_description')
    result = list()
    for info in query:
        info = {
            "date": info.date,
            "color": info.day_description.color,
            "description": info.day_description.description
        }
        result.append(info)
    return result


def get_all_faqs():
    """
    :return: all faqs
    """
    query = FAQ.objects.all()
    grouped = dict()
    for info in query:
        grouped.setdefault(info.category, []).append(info)
    return grouped


def search_faq(searchText):
    """
    :param search:
    :return: all faqs matched "search"
    """
    query = FAQ.objects.annotate(search=SearchVector('question', 'answer', config='french_unaccent')) \
        .filter(search=searchText)
    grouped = dict()
    for info in query:
        grouped.setdefault(info.category, []).append(info)
    return grouped


def get_op_route(auth_route_code):
    """
    :param auth_route_code:
    :return: op_route_code matched with auth_route_code
    """
    try:
        res = OPDictionary.objects.get(auth_route_code=auth_route_code).op_route_code
    except OPDictionary.DoesNotExist:
        return None
    return res


def get_op_routes_dict(key='auth_route_code', answer='route_type'):
    """
    :return: dict {key: answer}
    """
    op_program_dict = get_opprogram_list_for_select_input(to_dict=True)
    routes_dict = defaultdict(lambda: defaultdict(list))
    for op_key, op_answer, op_program in OPDictionary.objects.values_list(key, answer,
                                                                          'op_program_id'):
        routes_dict[op_program_dict[op_program]][op_key].append(op_answer)
    return routes_dict


def get_valid_time_period_date(date_list):
    """
    :param date_list:
    :return: (valid | not_valid, period_date_id)
    """
    valid_dates = TimePeriodDate.objects.values('date', 'id')
    period_date_valid = ''
    first_date = dt.fromisoformat(date_list[0])
    last_date = dt.fromisoformat(date_list[-1])
    for valid_date in valid_dates:
        if first_date < valid_date['date'] <= last_date:
            return False, ''
        else:
            if first_date >= valid_date['date']:
                period_date_valid = valid_date['id']
    return True, period_date_valid


def upload_xlsx_op_dictionary(xlsx_file, op_program_id):
    op_program = OPProgram.objects.get(id=op_program_id)
    upload_time = timezone.now()
    to_update = []
    to_create = []

    wb = load_workbook(xlsx_file, read_only=True)
    ws = wb.active
    key_indexes = [0, 5, 6, 7, 9]
    first_row = [ws['A2:P2'][0][x] for x in key_indexes]
    first_row_values = [cell.value for cell in first_row if cell.value is not None]

    if len(first_row_values) < 5:
        raise ValueError("Archivo con datos en blanco")
    for row in ws.iter_rows(min_row=1, max_col=16, values_only=True):
        if row[0] is None:
            break
        if row[4] == 'Vigente':
            auth_route_code = str(row[9])
            op_route_code = str(row[7]) + str(row[5])
            user_route_code = row[6]
            route_type = row[0]
            try:
                op_dict_obj = OPDictionary.objects.get(auth_route_code=auth_route_code, route_type=route_type,
                                                       op_program=op_program)
                op_dict_obj.auth_route_code = auth_route_code
                op_dict_obj.user_route_code = user_route_code
                op_dict_obj.route_type = route_type
                op_dict_obj.updated_at = upload_time
                to_update.append(op_dict_obj)

            except OPDictionary.DoesNotExist:
                to_create.append(OPDictionary(user_route_code=user_route_code, op_route_code=op_route_code,
                                              route_type=route_type, created_at=upload_time,
                                              updated_at=upload_time,
                                              auth_route_code=auth_route_code, op_program=op_program))
    wb.close()
    OPDictionary.objects.bulk_create(to_create)
    OPDictionary.objects.bulk_update(to_update,
                                     ['user_route_code', 'op_route_code', 'route_type', 'updated_at'])

    return {"created": len(to_create), "updated": len(to_update)}


def synchronize_op_program():
    """
    :return: info_dict
    """
    es_available_days = set(ESOPDataHelper().get_available_days())
    db_available_days = {str(op_program) for op_program in OPProgram.objects.all()}
    to_create = es_available_days - db_available_days
    to_delete = db_available_days - es_available_days

    OPProgram.objects.bulk_create([OPProgram(valid_from=day) for day in to_create])
    OPProgram.objects.filter(valid_from__in=to_delete).delete()
    return {'es_available_days': es_available_days, 'db_available_days': db_available_days, 'created': to_create,
            'deleted': to_delete}


def get_opprogram_list_for_select_input(to_dict=False):
    parser = _list_parser
    if to_dict:
        parser = _dict_parser
    return parser(OPProgram.objects.values_list('id').annotate(valid_from=Cast('valid_from', output_field=CharField())))


def get_periods_dict():
    """
    :return: dict with {id, [period list] }
    """
    time_period_date_ids = TimePeriodDate.objects.all().values_list('id', flat=True)
    time_period_date_dict = {}
    for date_id in time_period_date_ids:
        time_period_date_dict.update({date_id: get_timeperiod_list_for_select_input(filter_id=date_id)})
    return time_period_date_dict


class PermissionBuilder(object):

    def __init__(self):
        pass

    def create_permission_based_on_operators(self):
        """
        Create permissions and groups. It is used when localinfo initialize (see localinfo/apps.py)
        """
        group_names = []
        for operator in Operator.objects.all():
            permission, _ = GlobalPermission.objects.get_or_create(codename=operator.esId,
                                                                   defaults={'name': operator.name.lower()})
            group_names.append((operator.name.capitalize(), permission))

        global_group, _ = Group.objects.get_or_create(name=settings.GLOBAL_PERMISSION_GROUP_NAME)

        for group_name, permission in group_names:
            name = 'Datos de {0}'.format(group_name)
            group, _ = Group.objects.get_or_create(name=name)
            group.permissions.add(permission)
            global_group.permissions.add(permission)

        # create permission to see trip section and historical section
        trip_permission, _ = GlobalPermission.objects.get_or_create(codename='travel', defaults={'name': 'viajes'})
        general_permission, _ = GlobalPermission.objects.get_or_create(
            codename='globalstat', defaults={'name': 'estadísticas generales'})
        advance_group, _ = Group.objects.get_or_create(name='Sección viajes y estadísticas generales')
        advance_group.permissions.add(trip_permission, general_permission)

        # create permission to see storage section
        storage_permission, _ = GlobalPermission.objects.get_or_create(
            codename='storage', defaults={'name': 'almacenamiento'})
        storage_group, _ = Group.objects.get_or_create(name='Sección de almacenamiento')
        storage_group.permissions.add(storage_permission)

        # create permission to see bus station distribution section
        storage_permission, _ = GlobalPermission.objects.get_or_create(
            codename='validation', defaults={'name': 'validaciones'})
        storage_group, _ = Group.objects.get_or_create(name='Sección de validaciones')
        storage_group.permissions.add(storage_permission)

    def update_permission(self, new_operator_obj):
        """
        Modify permission if user change name or esId
        """
        old_operator = Operator.objects.get(id=new_operator_obj.id)
        permission = GlobalPermission.objects.get(codename=old_operator.esId)
        permission.codename = new_operator_obj.esId
        permission.name = new_operator_obj.name.lower()
        permission.save()

        group = Group.objects.get(name=old_operator.name.capitalize())
        group.name = new_operator_obj.name.capitalize()
        group.save()

    def add_permission(self, operator_obj):
        """
        create a new permission. It is used when user add new operator
        """
        permission, _ = GlobalPermission.objects.get_or_create(codename=operator_obj.esId,
                                                               name=operator_obj.name.lower())

        group, _ = Group.objects.get_or_create(name=operator_obj.name.capitalize())
        group.permissions.add(permission)

        global_group, _ = Group.objects.get_or_create(name=settings.GLOBAL_PERMISSION_GROUP_NAME)
        global_group.permissions.add(permission)

    def delete_permission(self, operator_obj):
        """
        delete permission related to operator obj
        """
        permission = GlobalPermission.objects.get(codename=operator_obj.esId)

        group = Group.objects.get(name=operator_obj.name.capitalize())
        group.permissions.remove(permission)
        group.delete()

        global_group = Group.objects.get(name=settings.GLOBAL_PERMISSION_GROUP_NAME)
        global_group.permissions.remove(permission)

        permission.delete()

    def get_valid_operator_id_list(self, user):
        """
        return list of operator esId field valid for user
        """
        answer = []
        for esId in Operator.objects.values_list('esId', flat=True):
            if user.has_perm('localinfo.{0}'.format(esId)):
                answer.append(esId)

        return answer
